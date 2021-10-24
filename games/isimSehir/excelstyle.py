from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side

thinBorder = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'))

# Background color presets
orangeFill = PatternFill(fgColor='FCD5B4', fill_type='solid')
oliveFill = PatternFill(fgColor='D8E4BC', fill_type='solid')
lightGrayFill = PatternFill(fgColor='F2F2F2', fill_type='solid')
darkGrayFill = PatternFill(fgColor='E2E2E2', fill_type='solid')
purpleFill = PatternFill(fgColor='E4DFEC', fill_type='solid')

centerIt = Alignment(horizontal='center', vertical='center')
boldFont = Font(bold=True)
