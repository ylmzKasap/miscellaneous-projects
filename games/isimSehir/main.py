import os
import random
import time
import datetime
import copy

import openpyxl
from openpyxl.utils import get_column_letter

from tablemaker import table_it
import playerclass
import whichtime
import excelstyle

players = []
gameCategories = []
playerNumber = categoryNumber = 1
letters = ['a', 'b', 'c', 'ç', 'd', 'e', 'f', 'g', 'ğ', 'h', 'ı', 'i',
           'j', 'k', 'l', 'm', 'n', 'o', 'ö', 'p', 'r', 's', 'ş', 't',
           'u', 'ü', 'v', 'y', 'z']

savePath = f'{os.getcwd()}\\Save Files'
if not os.path.exists(savePath):
    os.mkdir(savePath)

saveFolder = f"{len(os.listdir(savePath))}- {datetime.datetime.now().strftime('%d.%m.%Y')}"
savePath = f'{savePath}\\{saveFolder}'
os.makedirs(savePath)


def calculate_score(gameCategory, gameTour):
    answerToScore = []
    for gamePlayer in players:
        answerToScore.append(getattr(obj[gamePlayer], f'{gameCategory}_{gameTour}'))
    for gamePlayer in players:
        playerCategoryAnswer = getattr(obj[gamePlayer], f'{gameCategory}_{gameTour}')
        if playerCategoryAnswer == '':
            obj[gamePlayer].score_response(gameCategory, gameTour, 0)
        elif answerToScore.count('') == len(players) - 1:
            obj[gamePlayer].score_response(gameCategory, gameTour, 20)
        elif answerToScore.count(playerCategoryAnswer) > 1:
            obj[gamePlayer].score_response(gameCategory, gameTour, 5)
        else:
            obj[gamePlayer].score_response(gameCategory, gameTour, 10)


def get_answer_score_for_table(gameTour):
    for gamePlayer in players:
        for i in range(len(gameCategories)):
            obj[gamePlayer].table[i][gameTour] = (
                getattr(obj[gamePlayer], f'{gameCategories[i]}_{gameTour}')
                + ': '
                + str(getattr(obj[gamePlayer], f'{gameCategories[i]}_{gameTour}_score'))
                 )


def sum_tour_score(gameTour):
    for gamePlayer in players:
        playerRoundScore = 0
        for i in range(len(gameCategories)):
            playerRoundScore += getattr(obj[gamePlayer], f'{gameCategories[i]}_{gameTour}_score')
        try:
            obj[gamePlayer].table[-1][gameTour] = playerRoundScore
        except IndexError:
            obj[gamePlayer].table[-1].append(playerRoundScore)
        obj[gamePlayer].sum_tour(gameTour, playerRoundScore)


def save_to_excel():
    for gamePlayer in players:
        wBook = openpyxl.Workbook()
        sheet = wBook.active

        maxRow = len(obj[gamePlayer].table[0])
        maxColumn = len(obj[gamePlayer].table)

        # Create and color player name
        sheet['A1'] = f'{gamePlayer} Puan Tablosu'
        sheet.merge_cells(f'A1:{get_column_letter(maxColumn)}1')
        sheet['A1'].fill = excelstyle.orangeFill
        sheet['A1'].border = excelstyle.thinBorder

        # Style total score
        sumCell = f'{get_column_letter(maxColumn)}{maxRow + 2}'
        sheet[sumCell] = obj[gamePlayer].score
        sheet[sumCell].border = excelstyle.thinBorder
        sheet[sumCell].fill = excelstyle.purpleFill
        sheet[sumCell].alignment = excelstyle.centerIt
        sheet[sumCell].font = excelstyle.boldFont
        sheet.row_dimensions[maxRow + 2].height = 30

        for i in range(1, maxColumn + 1):
            # Color categories
            sheet[f'{get_column_letter(i)}2'].fill = excelstyle.oliveFill
            # Adjust column width
            sheet.column_dimensions[get_column_letter(i)].width = 20

        for i in range(1, 3):
            for j in range(1, maxColumn + 1):
                sheet[f'{get_column_letter(j)}{i}'].font = excelstyle.boldFont

        for i in range(1, maxRow + 2):
            # Adjust row height
            sheet.row_dimensions[i].height = 30

            for j in range(1, maxColumn + 1):
                styleCell = f'{get_column_letter(j)}{i}'
                # Adjust borders
                sheet[styleCell].border = excelstyle.thinBorder
                # Adjust alignment
                sheet[styleCell].alignment = excelstyle.centerIt
                # Wrap text
                alignment = copy.copy(sheet[styleCell].alignment)
                alignment.wrapText = True
                sheet[styleCell].alignment = alignment

            # Adjust row colors
            if i > 2 and i % 2 != 0:
                for j in range(1, maxColumn + 1):
                    sheet[f'{get_column_letter(j)}{i}'].fill = excelstyle.lightGrayFill
            elif i > 2 and i % 2 == 0:
                for j in range(1, maxColumn + 1):
                    sheet[f'{get_column_letter(j)}{i}'].fill = excelstyle.darkGrayFill

        # Write answers
        for columnIndex, column in enumerate(obj[gamePlayer].table):
            for rowIndex in range(len(column)):
                sheet.cell(column=columnIndex+1, row=rowIndex+2).value = obj[gamePlayer].table[columnIndex][rowIndex]

        try:
            wBook.save(f'{savePath}\\{gamePlayer}.xlsx')
        except PermissionError:
            pass


print('\nİsim şehir oyununa hoş geldiniz.\n', end=' ')

while True:
    print(f"\n{playerNumber}. oyuncunun ismini yazın."
          + " | '-' son kişiyi sil | 'q' isim sorgusunu bitir")
    name = input().title()

    if name in players:
        os.system('cls')
        print('\nBu isimde bir oyuncu zaten var. Başka bir isim girin.')
        continue

    if name == '':
        os.system('cls')
        print('\nBir isim girin.')
        continue

    if name == 'Q':
        if len(players) >= 2:
            break
        else:
            os.system('cls')
            print('\nOyunun başlaması için en az iki kişi olmalı. Bir oyuncu ismi girin.')
            continue

    try:
        os.system('cls')
        if name == '-':
            players.pop()
            playerNumber -= 1
            print(f"\nMevcut Oyuncular: {', '.join(players)}")
            continue

    except IndexError:
        os.system('cls')
        print('\nOlmayan şeyi nasıl sileyim yahu ?!')
        continue

    players.append(name)
    playerNumber += 1
    print(f"\nMevcut Oyuncular: {', '.join(players)}")

os.system('cls')

while True:
    print(f"\n{categoryNumber}. kategorinin ismini yazın."
          + " | '-' son kategoriyi sil | 'q' kategori sorgusunu bitir")
    category = input().title()

    if category in gameCategories:
        os.system('cls')
        print('\nBu isimde bir kategori zaten var. Başka bir kategori girin.')
        continue

    if category == '':
        os.system('cls')
        print('\nBir kategori girin.')
        continue

    if category == 'Q':
        if len(gameCategories) >= 2:
            break
        else:
            os.system('cls')
            print('\nOyunun başlaması için en az iki kategori olmalı. Bir kategori ismi girin.')
            continue

    try:
        os.system('cls')
        if category == '-':
            gameCategories.pop()
            categoryNumber -= 1
            print(f"\nMevcut Kategoriler: {', '.join(gameCategories)}")
            continue

    except IndexError:
        os.system('cls')
        print('\nOlmayan şeyi nasıl sileyim yahu ?!')
        continue

    gameCategories.append(category)
    categoryNumber += 1
    print(f"\nMevcut Kategoriler: {', '.join(gameCategories)}")

categoryNumbers = {category: index for index, category in enumerate(gameCategories)}

obj = {}
for player in players:
    obj[player] = playerclass.Contender(player, gameCategories)

for player in players:
    wb = openpyxl.Workbook()
    wb.save(f'{savePath}\\{player}.xlsx')

os.system('cls')
print('\nHarfler rastgele gelsin diyorsanız enterlayın gitsin.')
print('Harfleri kendiniz seçmek istiyorsanız \'b\' tuşuna basıp enterlayın.')

letterDecision = input().lower()
if letterDecision == 'b':
    letterDecision = 0
else:
    letterDecision = 1
os.system('cls')

tourTime = 0
while True:
    print('\nTur süresi kaç saniye olsun?')
    tourTime = input()
    try:
        tourTime = abs(int(tourTime))
        break
    except ValueError:
        os.system('cls')
        print('\nSaniye türünden geçerli bir sayı girin.')
os.system('cls')

tour = 1
while True:
    print('\nTuru başlatmak için entera basın.')
    changeTime = input()
    os.system('cls')

    if changeTime == 'süre':
        os.system('cls')
        while True:
            print('\nTur süresi kaç saniye olsun?')
            tourTime = input()
            try:
                tourTime = abs(int(tourTime))
                break
            except ValueError:
                os.system('cls')
                print('\nSaniye türünden geçerli bir sayı girin.')
        os.system('cls')
        print(f'\nYeni tur süresi: {whichtime.tell_time(tourTime)}')
        print('\nTuru başlatmak için entera basın.')
        changeTime = input()

    letterChosen = ''
    if letterDecision == 1:
        while True:
            letterChosen = random.choice(letters)

            if letterChosen == 'ğ' and len(letters) > 1:
                os.system('cls')
                print('\nTüh! \'Ğ\' geldi. Tekrar deneyin.')
                input()
                continue

            elif len(letters) == 1:
                os.system('cls')
                print('\nNasıl başardıysanız bütün harfleri tüketmişsiniz.'
                      + '\nYine de turu başlatmak için enterlayın.')
                input()
                letterDecision = 0
                os.system('cls')
                break

            os.system('cls')
            letters.remove(letterChosen)
            break

    countdown = tourTime
    while countdown > 0:
        try:
            if letterDecision == 1:
                print(f'\nSeçinlen harf: {letterChosen}')
            print('\nTur başladı!')
            print(f'\nTurun bitmesine son {whichtime.tell_time(countdown)}.')
            time.sleep(1)
            os.system('cls')
            countdown -= 1
        except KeyboardInterrupt:
            break

    os.system('cls')
    print('\nTur sona erdi. Cevaplara geçmek için enterlayın.')
    input()
    os.system('cls')

    categoryIndex = playerIndex = 0
    categoryAnswers = [
        [[player] for player in players] for category in gameCategories]
    goingBack = 0

    while categoryIndex < len(gameCategories):
        category = gameCategories[categoryIndex]

        if goingBack == 0:  # Player index should give the last player while going back
            playerIndex = 0
        elif goingBack == 1:
            goingBack = 0

        while playerIndex < len(players):
            player = players[playerIndex]
            print(f"\n{player} {category} için ne diyor?")
            playerAnswer = input().lower()
            playerAnswer = playerAnswer.strip()

            if playerAnswer == '-':
                try:
                    obj[players[playerIndex-1]].remove_answer(category, tour)
                    deleteOnAnswersList = categoryAnswers[categoryIndex][playerIndex-1]
                    if goingBack == 0 and len(deleteOnAnswersList) > 1:
                        deleteOnAnswersList.pop()
                        playerIndex -= 1
                    os.system('cls')
                    table_it(categoryAnswers[categoryIndex])
                    continue
                except AttributeError:
                    if categoryIndex > 0:
                        goingBack = 1
                        break
                    elif categoryIndex == 0:
                        os.system('cls')
                        continue

            obj[player].answer(category, playerAnswer)

            os.system('cls')
            categoryAnswers[categoryIndex][playerIndex].append(playerAnswer)
            table_it(categoryAnswers[categoryIndex])
            playerIndex += 1

        if goingBack == 1:
            categoryIndex -= 1
            playerIndex = len(players) - 1
            categoryAnswers[categoryIndex][playerIndex].pop()
            obj[players[playerIndex]].remove_answer(gameCategories[categoryIndex], tour)
            os.system('cls')
            table_it(categoryAnswers[categoryIndex])
            continue
        calculate_score(category, tour)
        categoryIndex += 1

    for category in gameCategories:
        for player in players:
            obj[player].score += getattr(obj[player], f'{category}_{tour}_score')  # Add to total score

    totalScores = {player: obj[player].score for player in players}
    sortedTotalScores = {
        k: v for k, v in sorted(totalScores.items(), reverse=True, key=lambda item: item[1])}
    scoresList = [[], []]
    for name, score in sortedTotalScores.items():
        scoresList[0].append(name)
        scoresList[1].append(score)

    get_answer_score_for_table(tour)
    sum_tour_score(tour)
    save_to_excel()

    os.system('cls')
    print('\nPuan tablosu: \n ')
    table_it(scoresList)

    info = 0
    while True:
        if info == 0:
            print('\nOyuncu ismi yazarak puan tablosunu görüntüleyebilirsiniz.')
            info = 1
        print(f"\nMevcut oyuncular: {', '.join(players)}")
        decision = input().title()
        if decision == '':
            break

        elif decision == '-':
            while decision != 'Q':
                os.system('cls')
                print('\nDüzeltme modu. \n\nCevabını düzenlemek istediğiniz kişinin ismini yazın.')
                table_it(scoresList)
                print('\nDüzeltme modundan çıkmak için \'q\' tuşuna basın.')
                playerNameToEdit = input().title()
                os.system('cls')

                if playerNameToEdit == 'Q':
                    decision = ' '
                    print('\nOyuncu ismi yazarak puan tablosunu görüntüleyebilirsiniz.')
                    table_it(scoresList)
                    break

                while playerNameToEdit not in players:
                    print(f"\nBöyle bir oyuncu yok. Mevcut oyuncular: {', '.join(players)}")
                    playerNameToEdit = input().title()
                    os.system('cls')

                print(f'\nOyuncu adı: {playerNameToEdit}')
                table_it(obj[playerNameToEdit].table)
                print('\nHangi kategorideki cevap değiştirilsin?')
                categoryToEdit = input().title()
                os.system('cls')

                while categoryToEdit not in gameCategories:
                    table_it(obj[playerNameToEdit].table)
                    print(f"\nBöyle bir kategori yok. \n\nMevcut kategoriler: {', '.join(gameCategories)}")
                    categoryToEdit = input().title()
                    os.system('cls')

                valueError = indexError = tourToEdit = 0
                while True:
                    try:
                        print(f'\nOyuncu adı: {playerNameToEdit}')
                        print(f'Kategori: {categoryToEdit}')
                        table_it(obj[playerNameToEdit].table)
                        print('\nHangi turdaki cevap değiştirilsin?')

                        if valueError == 1:
                            print('\nBir sayı girin.')
                            valueError = 0

                        if indexError == 1:
                            print(f'\nGeçerli bir tur sayısı girin. Şu anda {tour}. turdayız.')
                            indexError = 0

                        tourToEdit = int(input())

                    except ValueError:
                        valueError = 1
                        os.system('cls')
                        continue

                    if tourToEdit < 1 or tourToEdit > tour:
                        indexError = 1
                        os.system('cls')
                        continue
                    os.system('cls')
                    break

                print(f'\nOyuncu adı: {playerNameToEdit}')
                table_it(obj[playerNameToEdit].table)
                print('\n' + getattr(
                    obj[playerNameToEdit], f'{categoryToEdit}_{tourToEdit}')
                      + ' yerine ne yazılsın?')

                newAnswer = input().lower()
                newAnswer = newAnswer.strip()

                obj[playerNameToEdit].change_answer(categoryToEdit, tourToEdit, newAnswer)
                calculate_score(categoryToEdit, tourToEdit)
                get_answer_score_for_table(tourToEdit)

                for player in players:
                    obj[player].score -= getattr(obj[player], f'sum_{tourToEdit}')

                sum_tour_score(tourToEdit)

                for player in players:
                    obj[player].score += getattr(obj[player], f'sum_{tourToEdit}')

                totalScores = {player: obj[player].score for player in players}
                sortedTotalScores = {
                    k: v for k, v in sorted(totalScores.items(), reverse=True, key=lambda item: item[1])}
                scoresList = [[], []]
                for name, score in sortedTotalScores.items():
                    scoresList[0].append(name)
                    scoresList[1].append(score)
                os.system('cls')
                print()
                save_to_excel()
                table_it(scoresList)
                break
            continue

        elif decision == 'T':
            os.system('cls')
            table_it(scoresList)
            continue

        elif decision not in players:
            os.system('cls')
            table_it(scoresList)
            print(f'\n{decision} adında bir oyuncu yok.')
            continue

        os.system('cls')
        print(f'\n{decision} Puan Tablosu:')
        table_it(obj[decision].table)

    os.system('cls')
    tour += 1
