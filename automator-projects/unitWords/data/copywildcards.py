import importlib
import shutil
import os

import openpyxl

from . import searchinfo
from . import split_rows


def delete_folder_contents(folder):
    for filename in os.listdir(folder):
        file_path = os.path.join(folder, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print('Failed to delete %s. Reason: %s' % (file_path, e))


def copy_wildcards(projectPath):
    print("\nLocating files...")
    importlib.reload(searchinfo)
    if searchinfo.databaseDecision == "w":
        wb = openpyxl.load_workbook(f"{projectPath}\\data\\Wildcard Database.xlsx")
    elif searchinfo.databaseDecision == "v":
        wb = openpyxl.load_workbook(f"{projectPath}\\data\\Variable Database.xlsx")

    sheet = wb.active
    emptyColumns = []
    occupiedColumns = []
    for columnIndex, columns in enumerate(list(sheet.columns)):
        columnOccupied = 0
        for row in columns:
            if row.value is not None:
                columnOccupied = 1
                occupiedColumns.append(columnIndex + 1)
                break
        if columnOccupied == 0:
            emptyColumns.append(columnIndex + 1)

    maxRowLength = len(list(sheet.columns))

    foundFiles = []
    notFoundDict = {}
    locatedPaths = []
    for search in searchinfo.assignedSearches:
        columnsWords = list(sheet.columns)[search[2] - 1]
        words = [w.value for w in columnsWords if w.value is not None]
        extensions = search[1]
        foundFiles.append([search[2]])

        for word in words:
            extensionFound = 0
            for extension in extensions:
                for folderName, subFolder, fileNames in os.walk(search[0]):
                    for filename in fileNames:
                        if f"{word}.{extension}" == filename:
                            foundFiles[-1].append(filename)
                            locatedPaths.append(folderName + f"\\{filename}")
                            extensionFound = 1
                            break
                    if extensionFound == 1:
                        break
                if extensionFound == 1:
                    break
            if extensionFound == 0:
                if notFoundDict.setdefault(', '.join(extensions), [word]) != [word]:
                    notFoundDict[', '.join(extensions)].append(word)

    # Return if there are missing files
    os.system("cls")
    if len(notFoundDict) >= 1:
        textBlock = "\nSome files could not be found."
        for k, v in notFoundDict.items():
            textBlock += f"\n\n{k} files:"
            for wildcard in v:
                textBlock += f"\n{wildcard}"
        return textBlock, False

    # Delete old files and copy new files
    delete_folder_contents(f"{projectPath}\\search")
    for filePath in locatedPaths:
        shutil.copy(filePath, f"{projectPath}\\search")

    # Add files to the database
    if len(emptyColumns) >= 1:
        columnIndex = emptyColumns[0]
        del emptyColumns[0]
    else:
        columnIndex = maxRowLength + 1

    for wordsList in foundFiles:
        rowIndex = 1
        wordIndex = 1
        for i in range(len(list(sheet.columns)[wordsList[0] - 1])):
            if sheet.cell(column=wordsList[0], row=rowIndex).value is None:
                rowIndex += 1
            else:
                sheet.cell(column=columnIndex, row=rowIndex).value = wordsList[wordIndex]
                rowIndex += 1
                wordIndex += 1

        if len(emptyColumns) >= 1:
            columnIndex = emptyColumns[0]
            del emptyColumns[0]
        else:
            columnIndex += 1
            while columnIndex in occupiedColumns:
                columnIndex += 1

    if searchinfo.split_cells:
        wb = split_rows.split_rows(
            wb, searchinfo.splitPref[0], searchinfo.splitPref[1])
    try:
        wb.save(f"{projectPath}\\data\\Searched Database.xlsx")
    except PermissionError:
        return "File couldn't been saved is it is already open.", False
    return "Files are successfully copied.", True
