import openpyxl

from . import searchinfo


def split_rows(workbook, _range, repeat, ignoreSpace=True):
    """
    Takes a workbook object as an argument and returns a new workbook object
    with its cell values split in groups of 'range' and repeated 'repeat' times for each row.
    """

    sheet = workbook.active
    splitList = []
    rowsToSplit = list(sheet.rows)
    rangeCounter = 0

    for _row in rowsToSplit:
        for cell in _row:
            if ignoreSpace:
                if cell.value is None:
                    continue
            if rangeCounter % (_range * repeat) == 0:
                splitList.append([])
            splitList[-1].append(cell.value)
            rangeCounter += 1

    splitWb = openpyxl.Workbook()
    splitSheet = splitWb.active
    for i, _row in enumerate(splitList, 1):
        for j, cell in enumerate(_row, 1):
            splitSheet.cell(column=j, row=i).value = cell

    skip = searchinfo.skip_cells
    if skip:
        skippedList = []
        for _row in list(splitSheet.rows):
            skippedList.append([])
            for i in range(_range):
                skippedList[-1] += _row[i::3]

        skippedWb = openpyxl.Workbook()
        skippedSheet = skippedWb.active
        for i, _row in enumerate(skippedList, 1):
            for j, cell in enumerate(_row, 1):
                skippedSheet.cell(column=j, row=i).value = cell.value
        return skippedWb

    return splitWb
