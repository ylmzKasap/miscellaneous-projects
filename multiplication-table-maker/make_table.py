# A tool to create multiplication tables on Excel.
# Example: Enter 'make_table.py 10' to create a multiplication table of 10.

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
from pathlib import Path
import os
import sys

wb = openpyxl.Workbook()
sheet = wb.active

borderlands = Border(           # Thin border style
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'))

thickBorder = Border(           # Thick border style
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'),
    bottom=Side(style='thick'))

orangeFill = PatternFill(fgColor='FFBB44', fill_type='solid')    # Background color presets
yellowFill = PatternFill(fgColor='EEFF77', fill_type='solid')

centralize = Alignment(horizontal='center', vertical='center')   # Center align preset
boldFont = Font(bold=True)                                       # Bold preset

while True:
    try:
        number = sys.argv[1]
        number = int(number)
        if number >= 500:
            print('\nThis may take a while. Please wait...')
        break
    except ValueError:
        print('\nPlease enter a valid number to create a multiplication table.')
        sys.exit()

for i in range(1, number+1):
    sheet.cell(row=1, column=i+1).value = i                     # Top row input and styling
    sheet.cell(row=1, column=i+1).font = boldFont
    sheet.cell(row=1, column=i+1).alignment = centralize
    sheet.cell(row=1, column=i+1).border = thickBorder
    sheet.cell(row=1, column=i+1).fill = orangeFill

    sheet.cell(row=i+1, column=1).value = i                     # First column input and styling
    sheet.cell(row=i+1, column=1).font = boldFont
    sheet.cell(row=i+1, column=1).alignment = centralize
    sheet.cell(row=i+1, column=1).border = thickBorder
    sheet.cell(row=i+1, column=1).fill = orangeFill

for rows in range(2, sheet.max_row+1):
    for columns in range(2, sheet.max_column+1):
        sheet.cell(row=rows, column=columns).value = sheet.cell(  # Do-the-math-loop
            row=rows, column=1).value * sheet.cell(
            row=1, column=columns).value
        sheet.cell(row=rows, column=columns).alignment = centralize
        sheet.cell(row=rows, column=columns).border = borderlands
        if rows == columns:
            sheet.cell(row=rows, column=columns).fill = yellowFill

sheet.freeze_panes = 'B2'                                       # Freeze the first row and column

for i in range(1, sheet.max_column+1):
    sheet.column_dimensions[get_column_letter(i)].width = 6     # Set column width

for i in range(1, sheet.max_row+1):                             # Set row height
    sheet.row_dimensions[i].height = 20

try:
    fileName = f'multiplication table of {number}.xlsx'      # Save
    wb.save(fileName)
    print('\nThe file has been successfully saved.')
except PermissionError:
    print('\nProcess failed as the file is already open and cannot be overwritten.')
